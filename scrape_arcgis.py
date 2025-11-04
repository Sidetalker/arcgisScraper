"""CLI for exploring Summit County, CO short-term rental data via ArcGIS."""

from __future__ import annotations

import argparse
import csv
import getpass
import json
import logging
import math
import os
import re
import sys
from dataclasses import dataclass
from html import unescape
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, TextIO, Tuple

from arcgis.features import FeatureLayer
from arcgis.gis import GIS
from arcgis.geometry import Geometry
from arcgis.geometry.filters import intersects


# The hosted feature layer that powers the Summit County, CO Short-Term Rental
# public map. The layer exposes individual rental properties keyed by their
# Summit County schedule number (``Schno``) alongside permit metadata.
DEFAULT_LAYER_URL = (
    "https://services6.arcgis.com/dmNYNuTJZDtkcRJq/arcgis/rest/services/"
    "STR_Licenses_October_2025_public_view_layer/FeatureServer/0"
)

# Connecting to the county's ArcGIS Online organization makes it easy to reuse
# the same authenticated session if the user also needs to access other hosted
# content.
DEFAULT_PORTAL_URL = "https://summitcountyco.maps.arcgis.com"

# Summit County's feature services expect cross-domain requests to include a
# Referer header that matches the county's GIS hostname. Allow callers to
# override the header, but default to the county host so that queries work out
# of the box.
DEFAULT_REFERER = "https://experience.arcgis.com/experience/706a6886322445479abadb904db00bc0/"

NAME_FIELD = "Owner Name"
COMPANY_FIELD = "Company (Required if last name is not provided)"

DEFAULT_SHEETS_DOC_ID = "1kKuIBG3BQTKu3uiH3lcOg9o-fUJ79440FldeFO5gho0"
DEFAULT_COMPLEX_GID = "2088119676"
DEFAULT_OWNER_GID = "521649832"

BUSINESS_KEYWORDS = (
    " LLC",
    " L.L.C",
    " LLP",
    " L.L.P",
    " INC",
    " CO ",
    " COMPANY",
    " CORPORATION",
    " CORP",
    " LP",
    " L.P",
    " LLLP",
    " PLLC",
    " PC",
    " TRUST",
    " TR ",
    " FOUNDATION",
    " ASSOCIATES",
    " HOLDINGS",
    " ENTERPRISE",
    " ENTERPRISES",
    " PROPERTIES",
    " PROPERTY",
    " GROUP",
    " INVEST",
    " PARTNERSHIP",
    " PARTNERS",
    " LIVING TRUST",
    " REVOCABLE",
    " FAMILY",
    " MANAGEMENT",
    " FUND",
    " ESTATE",
    " LLC.",
    " LLC,",
)

SUFFIX_TOKENS = {"JR", "SR", "II", "III", "IV", "V"}

BR_SPLIT_RE = re.compile(r"<br\s*/?>", re.IGNORECASE)
TAG_RE = re.compile(r"<[^>]+>")
UNIT_RE = re.compile(r"UNIT\s+([A-Za-z0-9\\-]+)", re.IGNORECASE)
BLDG_RE = re.compile(r"\\bBLDG\\s+([A-Za-z0-9\\-]+)", re.IGNORECASE)


@dataclass
class QueryResult:
    """Container for aggregating paginated query responses."""

    template: Dict[str, Any]
    features: List[Dict[str, Any]]

    def to_dict(self) -> Dict[str, Any]:
        payload = dict(self.template)
        payload["features"] = list(self.features)
        payload.setdefault("exceededTransferLimit", False)
        return payload


def create_gis(
    portal_url: str,
    username: Optional[str],
    password: Optional[str],
    api_key: Optional[str],
    referer: Optional[str],
) -> GIS:
    """Authenticate against an ArcGIS portal."""

    logging.debug("TRACE: create_gis(referer='%s')", referer)

    if api_key:
        return GIS(portal_url, api_key=api_key, referer=referer)

    if username:
        if password is None:
            if sys.stdin.isatty():
                password = getpass.getpass(f"Password for {username}: ")
            else:  # pragma: no cover - non-interactive fallback
                raise RuntimeError("Password is required when providing --username")
        return GIS(portal_url, username, password, referer=referer)

    return GIS(portal_url, anonymous=True, referer=referer)


def resolve_layer(
    gis: GIS, layer_url: Optional[str], item_id: Optional[str], layer_index: int
) -> FeatureLayer:
    """Return the ArcGIS feature layer that should be queried."""

    logging.debug("TRACE: resolve_layer(layer_url='%s')", layer_url)

    if layer_url:
        return FeatureLayer(layer_url, gis=gis)

    if item_id:
        item = gis.content.get(item_id)
        if item is None:
            raise RuntimeError(f"Unable to find ArcGIS item with id '{item_id}'")

        try:
            # NB: The layer object inherits the parent's referer.
            return item.layers[layer_index]
        except (IndexError, AttributeError) as exc:  # pragma: no cover - defensive path
            raise RuntimeError(
                f"Item '{item_id}' does not expose a layer at index {layer_index}."
            ) from exc

    raise RuntimeError("Either --layer-url or --item-id must be provided")


def build_search_geometry(lat: float, lng: float, radius_m: float) -> Geometry:
    """Construct a WGS84 envelope around the requested coordinate."""

    meters_per_degree_lat = 111_320.0
    meters_per_degree_lng = meters_per_degree_lat * math.cos(math.radians(lat))

    if meters_per_degree_lng == 0:  # pragma: no cover - invalid latitude guard
        raise ValueError("Unable to compute longitude delta for the provided latitude")

    delta_lat = radius_m / meters_per_degree_lat
    delta_lng = radius_m / meters_per_degree_lng

    envelope = {
        "xmin": lng - delta_lng,
        "xmax": lng + delta_lng,
        "ymin": lat - delta_lat,
        "ymax": lat + delta_lat,
        "spatialReference": {"wkid": 4326},
    }
    return Geometry(envelope)


def _initial_page_size(layer: FeatureLayer, max_records: Optional[int]) -> int:
    default_size = getattr(getattr(layer, "properties", None), "maxRecordCount", 1000) or 1000
    if max_records is not None:
        return min(default_size, max_records)
    return default_size


def query_features(
    layer: FeatureLayer,
    geometry: Optional[Geometry],
    where: str,
    out_fields: str,
    return_geometry: bool,
    max_records: Optional[int],
) -> QueryResult:
    """Query the feature layer and page through the full response."""

    page_size = _initial_page_size(layer, max_records)
    offset = 0
    collected: List[Dict[str, Any]] = []
    template: Optional[Dict[str, Any]] = None

    while True:
        geometry_filter = intersects(geometry) if geometry is not None else None
        feature_set = layer.query(
            where=where,
            out_fields=out_fields,
            geometry_filter=geometry_filter,
            return_geometry=return_geometry,
            out_sr=4326,
            result_offset=offset,
            result_record_count=page_size,
        )
        page = feature_set.to_dict()

        if template is None:
            template = {k: v for k, v in page.items() if k != "features"}

        features = page.get("features", [])
        collected.extend(features)

        if max_records is not None and len(collected) >= max_records:
            collected = collected[:max_records]
            template["exceededTransferLimit"] = len(features) == page_size
            break

        if not features or len(features) < page_size:
            template["exceededTransferLimit"] = page.get("exceededTransferLimit", False)
            break

        offset += page_size

    assert template is not None  # pragma: no cover - template is set on first iteration
    return QueryResult(template=template, features=collected)


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Query a Summit County, CO short-term rental feature layer using the "
            "ArcGIS Python API and dump the raw JSON response for inspection."
        )
    )
    parser.add_argument("lat", type=float, nargs="?", help="Latitude in decimal degrees")
    parser.add_argument("lng", type=float, nargs="?", help="Longitude in decimal degrees")
    parser.add_argument(
        "-r",
        "--radius",
        type=float,
        default=400.0,
        help="Search radius in meters (default: 400)",
    )
    parser.add_argument(
        "--portal-url",
        default=DEFAULT_PORTAL_URL,
        help="ArcGIS portal URL to authenticate against (default: %(default)s)",
    )
    parser.add_argument(
        "--referer",
        default=os.getenv("ARCGIS_REFERER", DEFAULT_REFERER),
        help=(
            "Referer header to send with ArcGIS requests. Defaults to the Summit "
            "County GIS host so cross-domain queries succeed. Override when "
            "targeting other services."
        ),
    )
    parser.add_argument(
        "--api-key",
        default=os.getenv("ARCGIS_API_KEY"),
        help=(
            "ArcGIS API key for authentication (defaults to the ARCGIS_API_KEY "
            "environment variable if set)"
        ),
    )
    parser.add_argument("--username", help="ArcGIS username for authentication")
    parser.add_argument(
        "--password",
        help="Password for the supplied username (prompted if omitted)",
    )
    parser.add_argument(
        "--layer-url",
        default=DEFAULT_LAYER_URL,
        help="Feature layer URL to query (default: Summit County short-term rentals)",
    )
    parser.add_argument(
        "--item-id",
        help="ArcGIS item id containing the desired layer (used when --layer-url is omitted)",
    )
    parser.add_argument(
        "--layer-index",
        type=int,
        default=0,
        help="Layer index within the ArcGIS item (default: 0)",
    )
    parser.add_argument(
        "--where",
        default="1=1",
        help="Optional WHERE clause to further filter results",
    )
    parser.add_argument(
        "--out-fields",
        default="*",
        help="Comma-separated list of fields to return (default: all fields)",
    )
    parser.add_argument(
        "--max-records",
        type=int,
        help="Maximum number of records to return (defaults to service limit)",
    )
    parser.add_argument(
        "--no-geometry",
        dest="return_geometry",
        action="store_false",
        help="Omit geometry from the response payload",
    )
    parser.set_defaults(return_geometry=True)
    parser.add_argument("--output", help="Optional file path to save the JSON payload")
    parser.add_argument(
        "--owner-table",
        action="store_true",
        help=(
            "Format results as an owner contact CSV. This expects the layer to expose "
            "OwnerFullName and OwnerContactPublicMailingAddr fields, such as the "
            "PrISM parcel point views."
        ),
    )
    compliance_group = parser.add_argument_group("Compliance overlays")
    compliance_group.add_argument(
        "--zoning-layer-url",
        help=(
            "ArcGIS feature layer URL containing zoning attributes keyed by parcel schedule numbers."
        ),
    )
    compliance_group.add_argument(
        "--zoning-item-id",
        help="ArcGIS Online item id resolving to the zoning layer when a direct URL is not supplied.",
    )
    compliance_group.add_argument(
        "--zoning-layer-index",
        type=int,
        default=0,
        help="Layer index to use when resolving the zoning item id (default: 0).",
    )
    compliance_group.add_argument(
        "--zoning-join-field",
        default="SCHEDNUM",
        help="Field name in the zoning layer that stores the parcel schedule number (default: SCHEDNUM).",
    )
    compliance_group.add_argument(
        "--zoning-code-field",
        default="ZONE_CODE",
        help="Attribute containing the zoning district code to surface in exports (default: ZONE_CODE).",
    )
    compliance_group.add_argument(
        "--zoning-description-field",
        default="ZONE_DESC",
        help="Attribute containing the zoning description to surface in exports (default: ZONE_DESC).",
    )
    compliance_group.add_argument(
        "--zoning-where",
        help="Optional WHERE clause applied when querying the zoning layer.",
    )
    compliance_group.add_argument(
        "--land-use-layer-url",
        help=(
            "ArcGIS feature layer URL exposing land-use attributes keyed by parcel schedule numbers."
        ),
    )
    compliance_group.add_argument(
        "--land-use-item-id",
        help="ArcGIS Online item id resolving to the land-use layer when a direct URL is not supplied.",
    )
    compliance_group.add_argument(
        "--land-use-layer-index",
        type=int,
        default=0,
        help="Layer index to use when resolving the land-use item id (default: 0).",
    )
    compliance_group.add_argument(
        "--land-use-join-field",
        default="SCHEDNUM",
        help="Field name in the land-use layer that stores the parcel schedule number (default: SCHEDNUM).",
    )
    compliance_group.add_argument(
        "--land-use-code-field",
        default="LAND_USE",
        help="Attribute containing the land-use category code to surface in exports (default: LAND_USE).",
    )
    compliance_group.add_argument(
        "--land-use-description-field",
        default="LAND_DESC",
        help=(
            "Attribute containing the land-use description to surface in exports (default: LAND_DESC)."
        ),
    )
    compliance_group.add_argument(
        "--land-use-where",
        help="Optional WHERE clause applied when querying the land-use layer.",
    )
    parser.add_argument(
        "--all-subdivisions",
        action="store_true",
        help=(
            "Automatically enumerate unique SubdivisionName values within the search "
            "geometry and query each one sequentially. Useful when paired with "
            "--owner-table to dump every complex in the radius."
        ),
    )
    parser.add_argument(
        "--excel-output",
        help=(
            "Optional path to write a two-sheet Excel workbook (requires --owner-table). "
            "Sheet 1 lists properties by complex, sheet 2 groups them by owner with "
            "cross-sheet hyperlinks."
        ),
    )
    parser.add_argument(
        "--area",
        action="append",
        nargs=3,
        metavar=("LAT", "LNG", "RADIUS"),
        type=float,
        help=(
            "Additional search areas (latitude, longitude, radius in meters). "
            "May be supplied multiple times."
        ),
    )
    parser.add_argument(
        "--sheets-doc-id",
        default=DEFAULT_SHEETS_DOC_ID,
        help=(
            "Google Sheets document ID for direct hyperlinks (set to empty to disable)."
        ),
    )
    parser.add_argument(
        "--complex-gid",
        default=DEFAULT_COMPLEX_GID,
        help="GID of the complex sheet when building Google Sheets hyperlinks.",
    )
    parser.add_argument(
        "--owner-gid",
        default=DEFAULT_OWNER_GID,
        help="GID of the owner sheet when building Google Sheets hyperlinks.",
    )
    parser.add_argument(
        "--rewrite-xlsx",
        help=(
            "Path to an existing workbook whose Google Sheets hyperlinks should be rewritten. "
            "When supplied, no query is executed."
        ),
    )
    parser.add_argument(
        "--rewrite-output",
        help="Optional destination for the rewritten workbook (defaults to in-place overwrite).",
    )
    return parser.parse_args(argv)


def main():
    """Query the feature layer and print the resulting feature set."""

    args = parse_args(sys.argv[1:])

    if args.rewrite_xlsx:
        _rewrite_workbook_links(
            args.rewrite_xlsx,
            args.rewrite_output or args.rewrite_xlsx,
            args.sheets_doc_id,
            args.complex_gid,
            args.owner_gid,
        )
        return

    areas: List[Tuple[float, float, float]] = []
    if args.area:
        areas.extend([(lat, lng, radius) for lat, lng, radius in args.area])
    if args.lat is not None and args.lng is not None:
        areas.append((args.lat, args.lng, args.radius))

    if not areas:
        print(
            "At least one search area is required (lat/lng or --area) unless --rewrite-xlsx is provided",
            file=sys.stderr,
        )
        sys.exit(1)

    if args.excel_output and not args.owner_table:
        print("--excel-output requires --owner-table", file=sys.stderr)
        sys.exit(1)

    # The ArcGIS API for Python surfaces a lot of useful debugging information
    # via the root logger. Emit logs to stderr when producing CSV output so the
    # data stream remains clean.
    root_logger = logging.getLogger()
    root_logger.handlers.clear()
    root_logger.setLevel(logging.DEBUG)
    log_stream = sys.stderr if args.owner_table else sys.stdout
    handler = logging.StreamHandler(log_stream)
    handler.setLevel(logging.DEBUG)
    formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
    handler.setFormatter(formatter)
    root_logger.addHandler(handler)

    api_key = args.api_key
    referer = args.referer
    sheets_doc_id = (args.sheets_doc_id or "").strip()
    complex_gid = (args.complex_gid or "").strip()
    owner_gid = (args.owner_gid or "").strip()

    logging.debug("TRACE: main(referer='%s')", referer)

    try:
        gis = create_gis(args.portal_url, args.username, args.password, api_key, referer)
        layer = resolve_layer(gis, args.layer_url, args.item_id, args.layer_index)

        combined_features: List[Dict[str, Any]] = []
        seen_keys: Set[Any] = set()
        template: Optional[Dict[str, Any]] = None
        exceeded = False

        for lat, lng, radius in areas:
            search_geometry = build_search_geometry(lat, lng, radius)

            logging.debug("TRACE: main(query_features) area=(%s,%s,%s)", lat, lng, radius)
            if args.all_subdivisions:
                result = _query_all_subdivisions(
                    layer=layer,
                    geometry=search_geometry,
                    base_where=args.where,
                    out_fields=args.out_fields,
                    return_geometry=args.return_geometry,
                    max_records=args.max_records,
                )
            else:
                result = query_features(
                    layer,
                    search_geometry,
                    where=args.where,
                    out_fields=args.out_fields,
                    return_geometry=args.return_geometry,
                    max_records=args.max_records,
                )
            logging.debug("TRACE: main(query_features) -> done (area)")

            template = template or dict(result.template)
            exceeded = exceeded or result.template.get("exceededTransferLimit", False)
            _append_unique_features(combined_features, seen_keys, result.features)

            if args.max_records and len(combined_features) >= args.max_records:
                combined_features = combined_features[: args.max_records]
                exceeded = True
                break

        if template is None:
            template = {}
        template["exceededTransferLimit"] = exceeded
        result = QueryResult(template=template, features=combined_features)

        compliance_columns: List[str] = []
        schedule_numbers = _collect_schedule_numbers(result.features)

        if schedule_numbers and (args.zoning_layer_url or args.zoning_item_id):
            zoning_layer = resolve_layer(
                gis,
                args.zoning_layer_url,
                args.zoning_item_id,
                args.zoning_layer_index,
            )
            zoning_mapping = {
                "Zoning District": args.zoning_code_field,
                "Zoning Description": args.zoning_description_field,
            }
            zoning_overlay = _fetch_overlay_attributes(
                zoning_layer,
                schedule_numbers,
                join_field=args.zoning_join_field,
                where=args.zoning_where,
                value_fields=list(zoning_mapping.values()),
            )
            _apply_overlay_attributes(result.features, zoning_overlay, value_mapping=zoning_mapping)
            for column in zoning_mapping.keys():
                if column not in compliance_columns:
                    compliance_columns.append(column)

        if schedule_numbers and (args.land_use_layer_url or args.land_use_item_id):
            land_use_layer = resolve_layer(
                gis,
                args.land_use_layer_url,
                args.land_use_item_id,
                args.land_use_layer_index,
            )
            land_use_mapping = {
                "Land Use Category": args.land_use_code_field,
                "Land Use Description": args.land_use_description_field,
            }
            land_use_overlay = _fetch_overlay_attributes(
                land_use_layer,
                schedule_numbers,
                join_field=args.land_use_join_field,
                where=args.land_use_where,
                value_fields=list(land_use_mapping.values()),
            )
            _apply_overlay_attributes(result.features, land_use_overlay, value_mapping=land_use_mapping)
            for column in land_use_mapping.keys():
                if column not in compliance_columns:
                    compliance_columns.append(column)

        if args.owner_table:
            rows = _format_owner_table(result.features, extra_columns=compliance_columns)
            owners, property_to_owner = _build_owner_registry(rows)

            if sheets_doc_id and complex_gid and owner_gid:
                _apply_hyperlink_urls(
                    rows,
                    owners,
                    property_to_owner,
                    sheets_doc_id,
                    complex_gid,
                    owner_gid,
                )

            _emit_owner_table(rows, args.output, extra_columns=compliance_columns)
            if args.excel_output:
                _write_excel_workbook(
                    rows,
                    owners,
                    property_to_owner,
                    args.excel_output,
                    sheets_doc_id,
                    complex_gid,
                    owner_gid,
                )
        else:
            payload = json.dumps(result.to_dict(), indent=2)
            if args.output:
                with open(args.output, "w", encoding="utf-8") as fh:
                    fh.write(payload)
            print(payload)

    except Exception as exc:
        logging.debug("TRACE: main(exception: %s)", exc)
        print(f"Error querying ArcGIS feature layer: {exc}", file=sys.stderr)
        sys.exit(1)


def _query_all_subdivisions(
    layer: FeatureLayer,
    geometry: Geometry,
    base_where: str,
    out_fields: str,
    return_geometry: bool,
    max_records: Optional[int],
) -> QueryResult:
    filters = _collect_subdivision_filters(layer, geometry, base_where)
    logging.debug("TRACE: _query_all_subdivisions(filters=%d)", len(filters))

    if not filters:
        return query_features(
            layer,
            geometry,
            where=base_where,
            out_fields=out_fields,
            return_geometry=return_geometry,
            max_records=max_records,
        )

    aggregated: List[Dict[str, Any]] = []
    template: Optional[Dict[str, Any]] = None
    exceeded = False

    for label, clause in filters:
        where_clause = _combine_where(base_where, clause)
        logging.debug("TRACE: _query_all_subdivisions(query='%s')", where_clause)
        sub_result = query_features(
            layer,
            geometry,
            where=where_clause,
            out_fields=out_fields,
            return_geometry=return_geometry,
            max_records=None,
        )
        if template is None:
            template = dict(sub_result.template)
        aggregated.extend(sub_result.features)
        exceeded = exceeded or sub_result.template.get("exceededTransferLimit", False)

    if max_records is not None and len(aggregated) > max_records:
        aggregated = aggregated[:max_records]
        exceeded = True

    if template is None:
        template = {}
    template["exceededTransferLimit"] = exceeded
    return QueryResult(template=template, features=aggregated)


def _collect_subdivision_filters(
    layer: FeatureLayer,
    geometry: Geometry,
    base_where: str,
) -> List[tuple[str, str]]:
    result = query_features(
        layer,
        geometry,
        where=base_where,
        out_fields="SubdivisionName",
        return_geometry=False,
        max_records=None,
    )

    seen: set[str] = set()
    filters: List[tuple[str, str]] = []

    for feature in result.features:
        attrs = feature.get("attributes") or {}
        raw_name = (attrs.get("SubdivisionName") or "").strip()
        key = raw_name.upper() if raw_name else "__BLANK__"
        if key in seen:
            continue
        seen.add(key)

        if raw_name:
            clause = f"SubdivisionName = '{_escape_sql_literal(raw_name)}'"
            label = raw_name
        else:
            clause = "(SubdivisionName IS NULL OR SubdivisionName = '')"
            label = "Unspecified"

        filters.append((label, clause))

    filters.sort(key=lambda item: item[0].lower())
    return filters


def _combine_where(base_where: str, clause: str) -> str:
    base = (base_where or "").strip()
    if not clause or clause == "1=1":
        return base or clause or "1=1"
    if not base or base == "1=1":
        return clause
    return f"({base}) AND ({clause})"


def _escape_sql_literal(value: str) -> str:
    return value.replace("'", "''")


def _coerce_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def _chunked(values: Iterable[str], size: int) -> Iterable[List[str]]:
    batch: List[str] = []
    for value in values:
        batch.append(value)
        if len(batch) >= size:
            yield batch
            batch = []
    if batch:
        yield batch


def _build_in_clause(field: str, values: Sequence[str]) -> str:
    safe_values = [f"'{_escape_sql_literal(value)}'" for value in values if value]
    if not safe_values:
        return "1=0"
    joined = ", ".join(safe_values)
    return f"{field} IN ({joined})"


def _collect_schedule_numbers(features: List[Dict[str, Any]]) -> Set[str]:
    numbers: Set[str] = set()
    for feature in features:
        attrs = feature.get("attributes") or {}
        schedule_raw = attrs.get("PropertyScheduleText")
        if schedule_raw is None:
            continue
        schedule = _coerce_to_str(schedule_raw).strip()
        if schedule:
            numbers.add(schedule)
    return numbers


def _fetch_overlay_attributes(
    layer: FeatureLayer,
    schedule_numbers: Set[str],
    *,
    join_field: str,
    where: Optional[str],
    value_fields: Sequence[str],
) -> Dict[str, Dict[str, Any]]:
    if not schedule_numbers:
        return {}

    base_where = (where or "1=1").strip() or "1=1"
    all_fields = {join_field, *value_fields}
    out_fields = ",".join(sorted(all_fields))

    overlay_map: Dict[str, Dict[str, Any]] = {}
    chunk_size = 200

    for chunk in _chunked(sorted(schedule_numbers), chunk_size):
        clause = _build_in_clause(join_field, chunk)
        combined_where = _combine_where(base_where, clause)
        result = query_features(
            layer=layer,
            geometry=None,
            where=combined_where,
            out_fields=out_fields,
            return_geometry=False,
            max_records=None,
        )

        for feature in result.features:
            attrs = feature.get("attributes") or {}
            key_raw = attrs.get(join_field)
            key = _coerce_to_str(key_raw).strip()
            if not key:
                continue

            existing = overlay_map.setdefault(key, {})
            for field in value_fields:
                if field in attrs and attrs[field] is not None:
                    existing[field] = attrs[field]

    return overlay_map


def _apply_overlay_attributes(
    features: List[Dict[str, Any]],
    overlay_map: Dict[str, Dict[str, Any]],
    *,
    value_mapping: Dict[str, str],
) -> None:
    if not overlay_map:
        return

    for feature in features:
        attrs = feature.get("attributes")
        if not isinstance(attrs, dict):
            continue
        schedule = _coerce_to_str(attrs.get("PropertyScheduleText")).strip()
        if not schedule:
            continue

        overlay_values = overlay_map.get(schedule)
        if not overlay_values:
            continue

        for output_name, source_field in value_mapping.items():
            value = overlay_values.get(source_field)
            if value is None:
                continue
            attrs[output_name] = _coerce_to_str(value)


def _format_owner_table(
    features: List[Dict[str, Any]],
    extra_columns: Optional[Sequence[str]] = None,
) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for feature in features:
        attrs = feature.get("attributes", {})
        if not attrs:
            continue

        raw_names = _extract_owner_names(attrs)
        if not raw_names:
            fallback = attrs.get("OwnerFullName", "")
            raw_names = [fallback.strip()] if fallback else [""]

        address_line1, address_line2, city, state, postcode = _parse_owner_address(
            attrs.get("OwnerContactPublicMailingAddr", "")
        )
        complex_name = _normalize_complex_name(attrs)
        unit = _extract_unit(attrs)
        schedule_number = attrs.get("PropertyScheduleText", "") or ""
        detail_id = attrs.get("HC_RegistrationsOriginalCleaned") or schedule_number
        detail_url = (
            f"https://gis.summitcountyco.gov/map/DetailData.aspx?Schno={detail_id}"
            if detail_id
            else ""
        )
        physical_address = attrs.get("SitusAddress") or attrs.get("BriefPropertyDescription") or ""

        for raw_name in raw_names:
            first, middle, last, suffix, title, company = _split_owner_name(raw_name)
            owner_name = _aggregate_owner_name(first, middle, last, suffix, title, company)
            is_business = bool(company and company.strip())
            zip_code = postcode
            zip5 = zip_code.split("-")[0].strip() if zip_code else ""

            city_line = city or ""
            if city_line and state:
                city_line = f"{city_line}, {state}"
            elif state:
                city_line = state
            zip_for_line = zip_code or zip5
            if city_line and zip_for_line:
                city_line = f"{city_line} {zip_for_line}".strip()
            elif not city_line and zip_for_line:
                city_line = zip_for_line
            mailing_lines = [line for line in (address_line1, address_line2, city_line) if line]
            mailing_address = "\n".join(mailing_lines)

            row = {
                "Complex": complex_name,
                "Unit": unit,
                NAME_FIELD: owner_name,
                "Owner Link": "",
                "Business Owner?": "Yes" if is_business else "No",
                "Mailing Address": mailing_address,
                "Address Line 1": address_line1,
                "Address Line 2": address_line2,
                "City (Required)": city,
                "State": state,
                "Zip5": zip5,
                "Zip9": zip_code,
                "Subdivision": attrs.get("SubdivisionName", ""),
                "Schedule Number": schedule_number,
                "Public Detail URL": detail_url,
                "Physical Address": physical_address,
            }
            if extra_columns:
                for column in extra_columns:
                    row[column] = _coerce_to_str(attrs.get(column, ""))
            row.update(
                {
                    "First name": first,
                    "Middle": middle,
                    "Last Name": last,
                    "Suffix": suffix,
                    "Title": title,
                    "Company (Required if last name is not provided)": company,
                    "Original Zip": postcode,
                    "Comments": "",
                }
            )
            rows.append(row)

    rows.sort(key=lambda row: (row["Complex"].lower(), _unit_sort_key(row["Unit"])))
    return rows


IMPORTANT_COLUMNS = [
    "Complex",
    "Unit",
    NAME_FIELD,
    "Owner Link",
    "Business Owner?",
    "Mailing Address",
    "Address Line 1",
    "Address Line 2",
    "City (Required)",
    "State",
    "Zip5",
    "Zip9",
    "Subdivision",
    "Schedule Number",
    "Public Detail URL",
    "Physical Address",
]

SUPPLEMENTAL_COLUMNS = [
    "First name",
    "Middle",
    "Last Name",
    "Suffix",
    "Title",
    COMPANY_FIELD,
    "Original Zip",
    "Comments",
]


def _emit_owner_table(
    rows: List[Dict[str, str]],
    output_path: Optional[str],
    *,
    destination: Optional[TextIO] = None,
    extra_columns: Optional[Sequence[str]] = None,
) -> None:
    supplemental = list(SUPPLEMENTAL_COLUMNS)
    enriched = list(extra_columns or [])
    fieldnames = IMPORTANT_COLUMNS + enriched + supplemental

    close_stream = False
    if destination is None:
        if output_path:
            destination = open(output_path, "w", newline="", encoding="utf-8")
            close_stream = True
        else:
            destination = sys.stdout
    else:
        if output_path:
            raise ValueError("Specify either output_path or destination, not both.")

    writer = csv.DictWriter(destination, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

    if close_stream and destination is not None:
        destination.close()


def _build_owner_registry(
    rows: List[Dict[str, str]]
) -> tuple[List[Dict[str, Any]], Dict[int, Dict[str, Any]]]:
    owners: List[Dict[str, Any]] = []
    property_to_owner: Dict[int, Dict[str, Any]] = {}
    lookup: Dict[tuple, Dict[str, Any]] = {}

    for idx, row in enumerate(rows):
        key = _owner_key(row)
        owner_entry = lookup.get(key)
        if owner_entry is None:
            owner_entry = {
                "owner_id": f"OWN{len(owners) + 1:04d}",
                "first": row.get("First name", ""),
                "middle": row.get("Middle", ""),
                "last": row.get("Last Name", ""),
                "suffix": row.get("Suffix", ""),
                "company": row.get(COMPANY_FIELD, ""),
                "address1": row.get("Address Line 1", ""),
                "address2": row.get("Address Line 2", ""),
                "city": row.get("City (Required)", ""),
                "state": row.get("State", ""),
                "zip5": row.get("Zip5", ""),
                "zip9": row.get("Zip9", ""),
                "name": row.get(NAME_FIELD, ""),
                "mailing": row.get("Mailing Address", ""),
                "business": row.get("Business Owner?", ""),
                "properties": [],
            }
            owners.append(owner_entry)
            lookup[key] = owner_entry

        owner_entry["properties"].append(
            {
                "row_index": idx,
                "complex": row.get("Complex", ""),
                "unit": row.get("Unit", ""),
                "schedule": row.get("Schedule Number", ""),
            }
        )
        property_to_owner[idx] = owner_entry

    for owner in owners:
        owner["display_name"] = owner["name"].strip() or owner["owner_id"]
        owner["excel_row"] = None

    return owners, property_to_owner


def _apply_hyperlink_urls(
    rows: List[Dict[str, str]],
    owners: List[Dict[str, Any]],
    property_to_owner: Dict[int, Dict[str, Any]],
    doc_id: str,
    complex_gid: str,
    owner_gid: str,
) -> None:
    current_row = 2
    for owner in owners:
        owner["excel_row"] = current_row
        owner_url = (
            f"https://docs.google.com/spreadsheets/d/{doc_id}/edit#gid={owner_gid}&range=B{current_row}"
        )
        owner["owner_url"] = owner_url

        for prop in owner["properties"]:
            row_index = prop["row_index"]
            complex_row = row_index + 2
            complex_url = (
                f"https://docs.google.com/spreadsheets/d/{doc_id}/edit#gid={complex_gid}&range=A{complex_row}"
            )
            prop["owner_row"] = current_row
            prop["owner_url"] = owner_url
            prop["complex_row"] = complex_row
            prop["complex_url"] = complex_url

            label_parts = [prop.get("complex", ""), prop.get("unit", "")]
            if not prop.get("unit") and prop.get("schedule"):
                label_parts.append(prop["schedule"])
            label = " ".join(part for part in label_parts if part).strip() or prop.get("complex", "")
            prop["link_label"] = label

            rows[row_index]["Owner Link"] = (
                f'=HYPERLINK("{owner_url}", "{owner["owner_id"]}")'
            )

        current_row += len(owner["properties"])


def _append_unique_features(
    accumulator: List[Dict[str, Any]],
    seen: Set[Any],
    features: List[Dict[str, Any]],
) -> None:
    for feature in features:
        key = _feature_key(feature)
        if key in seen:
            continue
        seen.add(key)
        accumulator.append(feature)


def _feature_key(feature: Dict[str, Any]) -> Any:
    attrs = feature.get("attributes", {}) if isinstance(feature, dict) else {}
    schedule = attrs.get("PropertyScheduleText")
    parcel = attrs.get("HC_RegistrationsOriginalCleaned")
    object_id = attrs.get("OBJECTID")
    return (schedule, parcel, object_id)


def _aggregate_owner_name(
    first: str,
    middle: str,
    last: str,
    suffix: str,
    title: str,
    company: str,
) -> str:
    company = (company or "").strip()
    if company:
        return company

    parts: List[str] = []
    if title:
        parts.append(title.strip())
    if first:
        parts.append(first.strip())
    if middle:
        parts.append(middle.strip())
    if last:
        parts.append(last.strip())
    if suffix:
        suffix = suffix.strip()
        if parts:
            parts[-1] = f"{parts[-1]} {suffix}".strip()
        else:
            parts.append(suffix)

    return " ".join(part for part in parts if part).strip()


def _owner_key(row: Dict[str, str]) -> tuple:
    company = (row.get(COMPANY_FIELD) or "").strip().upper()
    parts = (
        company,
        (row.get("First name") or "").strip().upper(),
        (row.get("Middle") or "").strip().upper(),
        (row.get("Last Name") or "").strip().upper(),
        (row.get("Suffix") or "").strip().upper(),
    )
    return parts


def _write_excel_workbook(
    rows: List[Dict[str, str]],
    owners: List[Dict[str, Any]],
    property_to_owner: Dict[int, Dict[str, Any]],
    path: str,
    doc_id: str,
    complex_gid: str,
    owner_gid: str,
) -> None:
    if not rows:
        raise RuntimeError("No results available to write to the Excel workbook.")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError("openpyxl is required for --excel-output. Install via 'pip install openpyxl'.") from exc

    wb = Workbook()
    ws_complex = wb.active
    ws_complex.title = "By Complex"
    ws_owner = wb.create_sheet("By Owner")

    base_headers = list(rows[0].keys())
    complex_headers = base_headers + ["Owner ID", "Owner Link"]
    ws_complex.append(complex_headers)

    if any(owner.get("excel_row") is None for owner in owners):
        row_counter = 2
        for owner in owners:
            owner["excel_row"] = row_counter
            row_counter += len(owner["properties"])

    owner_link_index = base_headers.index("Owner Link") + 1 if "Owner Link" in base_headers else None
    mailing_index = base_headers.index("Mailing Address") + 1 if "Mailing Address" in base_headers else None

    property_row_numbers: Dict[int, int] = {}

    for idx, row in enumerate(rows):
        owner_entry = property_to_owner[idx]
        excel_row = idx + 2
        property_row_numbers[idx] = excel_row

        owner_id = owner_entry["owner_id"]
        owner_ref_default = f"'By Owner'!B{owner_entry['excel_row']}"
        owner_link_value = rows[idx].get("Owner Link") or owner_ref_default
        values = [row.get(header, "") for header in base_headers] + [owner_id, owner_link_value]
        ws_complex.append(values)

        link_cell = ws_complex.cell(row=excel_row, column=len(complex_headers))
        link_cell.value = owner_link_value

        if owner_link_index:
            owner_link_cell = ws_complex.cell(row=excel_row, column=owner_link_index)
            owner_link_cell.value = owner_link_value

        if mailing_index:
            ws_complex.cell(row=excel_row, column=mailing_index).alignment = Alignment(wrap_text=True)

    for col_idx, header in enumerate(complex_headers, start=1):
        ws_complex.cell(row=1, column=col_idx).value = header

    owner_headers = [
        "Owner ID",
        NAME_FIELD,
        "Business Owner?",
        "Mailing Address",
        "Address Line 1",
        "Address Line 2",
        "City (Required)",
        "State",
        "Zip5",
        "Zip9",
        COMPANY_FIELD,
        "First name",
        "Middle",
        "Last Name",
        "Suffix",
        "Property Index",
        "Property Complex",
        "Property Unit",
        "Schedule Number",
        "Complex Sheet Link",
    ]
    ws_owner.append(owner_headers)

    property_link_index = owner_headers.index("Complex Sheet Link") + 1
    alignment_wrap = Alignment(wrap_text=True)

    current_row = 2
    for owner in owners:
        for prop_index, prop in enumerate(owner["properties"], start=1):
            excel_row = current_row
            current_row += 1

            link_row = property_row_numbers[prop["row_index"]]
            complex_ref = f"'By Complex'!A{link_row}"
            if prop.get("complex_url"):
                complex_formula = (
                    f'=HYPERLINK("{prop["complex_url"]}", "{_excel_escape(prop.get("link_label", complex_ref))}")'
                )
            else:
                complex_formula = complex_ref

            row_values = [
                owner["owner_id"],
                owner["name"],
                owner["business"],
                owner["mailing"],
                owner["address1"],
                owner["address2"],
                owner["city"],
                owner["state"],
                owner["zip5"],
                owner["zip9"],
                owner["company"],
                owner["first"],
                owner["middle"],
                owner["last"],
                owner["suffix"],
                prop_index,
                prop["complex"],
                prop["unit"],
                prop.get("schedule", ""),
                complex_formula,
            ]

            ws_owner.append(row_values)

            addr_cell = ws_owner.cell(row=excel_row, column=4)
            addr_cell.alignment = alignment_wrap

            prop_cell = ws_owner.cell(row=excel_row, column=property_link_index)
            prop_cell.value = complex_formula

    for col_idx, header in enumerate(owner_headers, start=1):
        ws_owner.cell(row=1, column=col_idx).value = header

    _write_instructions_sheet(wb, doc_id, complex_gid, owner_gid)
    wb.save(path)


def _write_instructions_sheet(wb, doc_id: str, complex_gid: str, owner_gid: str) -> None:
    note = (
        "Hyperlinks in the 'Owner Link' and 'Complex Sheet Link' columns use the Google Sheets "
        f"document id {doc_id or '<unset>'} with gids {complex_gid or '<unset>'}/{owner_gid or '<unset>'}. "
        "Adjust the --sheets-doc-id/--complex-gid/--owner-gid flags when exporting or run the script with "
        "--rewrite-xlsx to retarget links after moving this workbook."
    )

    lines = [
        "Workbook Notes:",
        note,
        "Cells show friendly text (e.g. OWN0001) but link to the corresponding tab/range in Google Sheets.",
        "Use Google Sheets' 'Open link' action or Excel/Numbers to follow them directly.",
    ]

    if "Instructions" in wb.sheetnames:
        sheet = wb["Instructions"]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = wb.create_sheet("Instructions", 0)

    for idx, line in enumerate(lines, start=1):
        sheet.cell(row=idx, column=1).value = line
    sheet.column_dimensions["A"].width = 110


def _rewrite_workbook_links(
    src_path: str,
    dest_path: str,
    doc_id: str,
    complex_gid: str,
    owner_gid: str,
) -> None:
    if not (doc_id and complex_gid and owner_gid):
        raise RuntimeError("--sheets-doc-id, --complex-gid, and --owner-gid are required when rewriting hyperlinks.")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError("openpyxl is required for --rewrite-xlsx. Install via 'pip install openpyxl'.") from exc

    wb = load_workbook(src_path)

    try:
        ws_complex = wb["By Complex"]
        ws_owner = wb["By Owner"]
    except KeyError as exc:
        raise RuntimeError("Workbook must contain 'By Complex' and 'By Owner' sheets to rewrite links.") from exc

    owner_link_col = _find_header_column(ws_complex, "Owner Link")
    owner_id_col = _find_header_column(ws_complex, "Owner ID")
    complex_link_col = _find_header_column(ws_owner, "Complex Sheet Link")
    prop_complex_col = _find_header_column(ws_owner, "Property Complex")
    prop_unit_col = _find_header_column(ws_owner, "Property Unit")
    schedule_col = _find_header_column(ws_owner, "Schedule Number")

    for row in range(2, ws_complex.max_row + 1):
        cell = ws_complex.cell(row=row, column=owner_link_col)
        value = cell.value
        if not isinstance(value, str) or not value.strip():
            continue
        owner_id = ws_complex.cell(row=row, column=owner_id_col).value or ""
        cell_range, label = _extract_range_and_label(value, default_label=owner_id)
        if not cell_range:
            continue
        cell.value = _build_hyperlink_formula(doc_id, owner_gid, cell_range, label)

    for row in range(2, ws_owner.max_row + 1):
        cell = ws_owner.cell(row=row, column=complex_link_col)
        value = cell.value
        if not isinstance(value, str) or not value.strip():
            continue
        default_label_parts = [ws_owner.cell(row=row, column=prop_complex_col).value or ""]
        unit_val = ws_owner.cell(row=row, column=prop_unit_col).value if prop_unit_col else ""
        if unit_val:
            default_label_parts.append(str(unit_val))
        sched_val = ws_owner.cell(row=row, column=schedule_col).value if schedule_col else ""
        if sched_val and not unit_val:
            default_label_parts.append(str(sched_val))
        default_label = " ".join(str(part) for part in default_label_parts if part).strip()
        cell_range, label = _extract_range_and_label(value, default_label=default_label)
        if not cell_range:
            continue
        cell.value = _build_hyperlink_formula(doc_id, complex_gid, cell_range, label)

    _write_instructions_sheet(wb, doc_id, complex_gid, owner_gid)
    wb.save(dest_path)


HYPERLINK_RE = re.compile(
    r'=HYPERLINK\("https://docs\.google\.com/spreadsheets/d/[^/]+/edit#gid=\d+&range=([A-Z]+\d+)"\s*,\s*"((?:""|[^"])*)"\)'
)
CELL_REF_RE = re.compile(r"'[^']+'!([A-Z]+\d+)")


def _extract_range_and_label(value: str, default_label: str = "") -> tuple[Optional[str], str]:
    value = value.strip()
    match = HYPERLINK_RE.match(value)
    if match:
        cell_range = match.group(1)
        label = match.group(2).replace('""', '"')
        return cell_range, label

    ref_match = CELL_REF_RE.search(value)
    if ref_match:
        return ref_match.group(1), default_label

    return None, default_label


def _build_hyperlink_formula(doc_id: str, gid: str, cell_range: str, label: str) -> str:
    escaped_label = _excel_escape(label or cell_range)
    return (
        f'=HYPERLINK("https://docs.google.com/spreadsheets/d/{doc_id}/edit#gid={gid}&range={cell_range}", '
        f'"{escaped_label}")'
    )


def _find_header_column(sheet, header: str) -> int:
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if isinstance(value, str) and value.strip() == header:
            return col
    raise RuntimeError(f"Unable to locate header '{header}' in sheet '{sheet.title}'.")


def _excel_escape(label: str) -> str:
    return label.replace('"', '""')


def _extract_owner_names(attrs: Dict[str, Any]) -> List[str]:
    html_names = attrs.get("OwnerNamesPublicHTML")
    if not html_names:
        return []

    decoded = unescape(html_names)
    parts = [TAG_RE.sub("", part).strip() for part in BR_SPLIT_RE.split(decoded)]
    return [part for part in parts if part]


def _parse_owner_address(raw: str) -> tuple[str, str, str, str, str]:
    if not raw:
        return "", "", "", "", ""

    decoded = unescape(raw)
    segments = [segment.strip() for segment in decoded.split("|") if segment.strip()]
    if not segments:
        return "", "", "", "", ""

    line1 = segments[0]
    line2 = ""
    city = state = postcode = ""
    city_state = ""

    if len(segments) == 2:
        city_state = segments[1]
    elif len(segments) >= 3:
        line2 = " ".join(segments[1:-1])
        city_state = segments[-1]

    if city_state:
        if "," in city_state:
            city_part, rest = city_state.split(",", 1)
            city = city_part.strip().title()
            rest = rest.strip()
            if rest:
                tokens = rest.split()
                if tokens:
                    state = tokens[0].upper()
                    postcode = " ".join(tokens[1:]).strip()
        else:
            city = city_state.title()

    return line1, line2, city, state, postcode


def _normalize_complex_name(attrs: Dict[str, Any]) -> str:
    subdivision = (attrs.get("SubdivisionName") or "").title().strip()
    if subdivision:
        for suffix in (" Condo", " Condos", " Condominiums", " Townhomes", " Townhome", " Pud", " Filing", " Phase"):
            if subdivision.endswith(suffix):
                subdivision = subdivision[: -len(suffix)].strip()
        replacements = {
            "Mountain Thunder Lodge": "Mountain Thunder",
        }
        return replacements.get(subdivision, subdivision)

    situs = attrs.get("SitusAddress", "")
    if not situs:
        return ""

    parts = situs.split()
    if parts and parts[0].isdigit():
        parts = parts[1:]

    trimmed: List[str] = []
    for part in parts:
        upper = part.upper()
        if upper in {"UNIT", "BLDG", "BUILDING"}:
            break
        trimmed.append(part)

    if trimmed:
        return " ".join(trimmed).title()
    return situs


def _extract_unit(attrs: Dict[str, Any]) -> str:
    for text in (attrs.get("BriefPropertyDescription"), attrs.get("SitusAddress")):
        if not text:
            continue
        match = UNIT_RE.search(str(text))
        if match:
            return match.group(1)
    for text in (attrs.get("BriefPropertyDescription"), attrs.get("SitusAddress")):
        if not text:
            continue
        match = BLDG_RE.search(str(text))
        if match:
            return match.group(1)
    return ""


def _split_owner_name(raw_name: str) -> tuple[str, str, str, str, str, str]:
    clean = raw_name.strip().strip(",")
    if not clean:
        return "", "", "", "", "", ""

    clean = clean.replace("  ", " ")
    upper = clean.upper()
    if any(keyword in upper for keyword in BUSINESS_KEYWORDS):
        return "", "", "", "", "", clean

    tokens = clean.replace(".", "").split()
    if not tokens:
        return "", "", "", "", "", ""

    suffix = ""
    if tokens[-1].upper() in SUFFIX_TOKENS:
        suffix = tokens.pop(-1)

    if not tokens:
        return "", "", "", suffix, "", ""

    if len(tokens) == 1:
        return "", "", tokens[0].title(), suffix, "", ""

    first_middle = tokens[:-1]
    last = tokens[-1].title()

    if any(token.upper() in {"&", "AND"} for token in first_middle):
        first = " ".join(token.title() for token in first_middle)
        middle = ""
    else:
        first = first_middle[0].title()
        middle = " ".join(token.title() for token in first_middle[1:])

    return first, middle, last, suffix, "", ""


def _unit_sort_key(unit: str) -> tuple[int, str]:
    if not unit:
        return (1, "")
    try:
        value = float(unit)
    except ValueError:
        return (0, unit.lower())
    return (0, f"{value:012.4f}")


if __name__ == "__main__":
    main()
COMPANY_FIELD = "Company (Required if last name is not provided)"
NAME_FIELD = "Owner Name"
